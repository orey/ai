#-------------------------------------------------------------------
# The objective is to decode Excel spreadsheet and extract text
# Each tab is a separate txt file.
# The program proposes a MD format and a CSV format
# APIs:
# - get_document_text
# - treat_text_with_contents
#-------------------------------------------------------------------
from openpyxl import load_workbook

import sys
sys.path.append('.')

from tools10 import myprint, is_zip, ensureFile, create_text_file

# Some spreadsheets have worksheets with a number of rows that is superior
# to this limit. We decided to cap to ROWS_LIMIT
ROWS_LIMIT = 5000


#---------------------------------------------------------- clean_spreadsheet_name
def clean_spreadsheet_name(name: str)-> str:
    clean = ""
    for x in name:
        if x.isalnum():
            clean += x
        else:
            clean += "_"
    return clean


#---------------------------------------------------------- extract xlsx
def treat_xlsx_tabs(source_file, #completefilename
               target_file, #completefilename
               format="MD", # or ",",";" or "|"
               verbose=False):
    '''
    This method keeps all the empty lines.
    I don't know what it gives on merges cells.
    It can probably extract the pricing lists quite efficiently
    return False in case of trouble, and sheets_status in case of treatment
    '''
    try:
        # 1- Test if the file is genuine
        if not is_zip(source_file):
            myprint(f"| XLSX | Not a regular XLSX | File {source_file}")
            FILES_NOT_OK.add([source_file])
            return False
        # 2- Loading the file
        wb = load_workbook(source_file, read_only=True, data_only=True)
        sheets_status = {} # { sheet_name, True/False }
        # 3- Looping on sheets
        for sheet_name in wb.sheetnames:
            # see if we have already processed the sheet
            if format == "MD":
                target = target_file.replace(".xlsx","") + "_[" + clean_spreadsheet_name(sheet_name) + "].xlsx.md"
            else:
                target = target_file.replace(".xlsx","") + "_[" + clean_spreadsheet_name(sheet_name) + "].xlsx.csv"
            if ensureFile(target):
                myprint(f"| XLSX | File already processed: {target}")
                continue
            text = ""
            ws = wb[sheet_name]
            #text += f"Sheet named '{sheet_name}' content hereafter:\n\n"
            myprint(f"| XLSX | Sheet: [{sheet_name}] with rows/columns = {ws.max_row}/{ws.max_column}", verbose)
            if ws.max_row > ROWS_LIMIT:
                myprint(f"| XSLX | Sheet: [{sheet_name}] too large. Only {ROWS_LIMIT} rows will be processed", verbose)
            # 4- Main loop on rows
            countrows = 0
            for row in ws.iter_rows():
                countrows += 1
                if countrows > ROWS_LIMIT:
                    myprint(f"| XLSX | === Sheet: [{sheet_name}] too large. We have processed the {ROWS_LIMIT} rows only", verbose)
                    break
                if verbose:
                    print(".", end="", flush=True)
                row_data = []
                for cell in row:
                    # Skip empty cells if needed
                    value = cell.value if cell.value is not None else ""
                    row_data.append(str(value))

                # 5- Process each row (myprint, save, etc.)
                #myprint(row_data)
                if format == "MD":
                    text += "| " + " | ".join(row_data) + ' |' + '\n'
                else:
                    theformat = ";"
                    if format in [',','|']:
                        theformat = format
                    text += theformat.join(row_data) + '\n'
            if verbose:
                print("", flush=True) # because the "."
            # 6- Creating one file per spreadsheet
            if create_text_file(target, text, verbose=verbose, prefix="| XLSX | "):
                myprint(f"| XLSX | File for [{sheet_name}] created: {target}")
                sheets_status[sheet_name] = True
            else:
                sheets_status[sheet_name] = False
        wb.close()
        return sheets_status
    except Exception as e:
        myprint(e)
        return False

    
#------------------------------------------------------------------main
if __name__ == "__main__":
    nb = len(sys.argv)
    if nb == 1:
        sst1 = treat_xlsx_tabs("./data/Book1.xlsx", #completefilename
                               "./data/OutBook.xlsx", #completefilename
                               format="MD", # or ",",";" or "|"
                               verbose=True)
        print(sst1)
        sst2 = treat_xlsx_tabs("./data/Book1.xlsx", #completefilename
                               "./data/OutBook.xlsx", #completefilename
                               format=";", # or ",",";" or "|"
                               verbose=True)
        print(sst2)
    else:
        if not ensureFile(sys.argv[1]):
            print(f"Argument problem: {sys.argv[1]}. Exiting.")
        else:
            sst = treat_xlsx_tabs(sys.argv[1], #completefilename
                                  sys.argv[1], #completefilename
                                  format=";", # or ",",";" or "|"
                                  verbose=True)
            print(sst)


    
